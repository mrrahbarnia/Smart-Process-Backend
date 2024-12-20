import logging
import asyncio
import sqlalchemy as sa
import sqlalchemy.orm as so

from io import BytesIO
from openpyxl import load_workbook # type: ignore
from fastapi import UploadFile
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession, AsyncEngine, async_sessionmaker
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert as postgres_upsert

from src.admin import schemas
from src.admin import exceptions
from src.admin.models import Guaranty
from src.admin.types import AdminProductDetailResponse, ExcelEntityTypes
from src.admin.utils import (
    validate_images_and_return_unique_image_names,
    create_unique_excel_name
)
from src.pagination import paginate
from src.products.types import (
    CategoryId,
    BrandId,
    ProductId,
    CommentId,
    SerialNumber
)
from src.products.models import (
    Brand,
    Category,
    Attribute,
    CategoryAttribute,
    Product,
    ProductImage,
    AttributeValue,
    Comment
)
from src.s3.utils import upload_to_s3, delete_from_s3
from src.tickets.models import Ticket
from src.tickets.types import TicketId
from src.articles.models import (
    Tag,
    ArticleTag,
    Article,
    ArticleImage,
    GlossaryTerm,
    ArticleComment
)
from src.articles.types import ArticleId, GlossaryId, ArticleCommentId
from src.articles.exceptions import ArticleNotFound

logger = logging.getLogger("admin")

# ==================== Brand service ==================== #

async def create_brand(
        payload: schemas.Brand,
        session: async_sessionmaker[AsyncSession],
        redis: Redis
) -> None:
    query = sa.insert(Brand).values(
        {
            Brand.name: payload.name,
            Brand.slug: payload.slug,
            Brand.description: payload.description
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
        await redis.delete("brand-list")
    except IntegrityError as ex:
        logger.warning(ex)
        if "uq_brands_name" in str(ex):
            raise exceptions.UniqueConstraintBrandName


async def activate_brand(
        slug: str,
        session: async_sessionmaker[AsyncSession],
        redis: Redis
) -> None:
    query = sa.update(Brand).where(Brand.slug==slug).values(
        {
            Brand.is_active: True
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)
    await redis.delete("brand-list")


async def update_brand_by_slug(
        brand_slug: str,
        session: async_sessionmaker[AsyncSession],
        payload: schemas.Brand,
        redis: Redis
) -> None:
    query = sa.update(Brand).values(
        {
            Brand.name: payload.name,
            Brand.description: payload.description,
            Brand.slug: payload.slug
        }
    ).where(Brand.slug==brand_slug).returning(Brand.id)
    try:
        async with session.begin() as conn:
            result: BrandId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.BrandNotFound
            await redis.delete("brand-list")
    except exceptions.BrandNotFound as ex:
        logger.warning(ex)
        raise exceptions.BrandNotFound
    except IntegrityError as ex:
        logger.warning(ex)


async def deactivate_brand(
        slug: str,
        session: async_sessionmaker[AsyncSession],
        redis: Redis
) -> None:
    query = sa.update(Brand).where(Brand.slug==slug).values(
        {
            Brand.is_active: False
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except Exception as ex:
        logger.warning(ex)
    await redis.delete("brand-list")


async def delete_brand(
        slug: str,
        session: async_sessionmaker[AsyncSession],
        redis: Redis
) -> None:
    query = sa.delete(Brand).where(Brand.slug==slug)
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except Exception as ex:
        logger.warning(ex)
    await redis.delete("brand-list")


async def all_brands(
        engine: AsyncEngine, limit: int, offset: int
) -> dict | None:
    query = sa.select(
        Brand.name, Brand.slug, Brand.description, Brand.is_active
    ).order_by(Brand.created_at.desc())
    return await paginate(
        engine=engine, query=query, limit=limit, offset=offset
    )

# ==================== Category service ==================== #

async def add_category(
        session: async_sessionmaker[AsyncSession],
        payload: schemas.Category,
        redis: Redis
) -> None:
    if payload.parent_category_name:
        parent_query = sa.select(Category.id).where(Category.name==payload.parent_category_name)
        try:
            async with session.begin() as conn:
                parent_category_id: CategoryId | None = await conn.scalar(parent_query)
                if parent_category_id is None:
                    raise exceptions.InvalidParentCategoryName
                query = sa.insert(Category).values(
                    {
                        Category.name: payload.name,
                        Category.description: payload.description,
                        Category.parent_id: parent_category_id
                    }
                )
                await conn.execute(query)
        except exceptions.InvalidParentCategoryName as ex:
            logger.warning(ex)
            raise exceptions.InvalidParentCategoryName
        except IntegrityError as ex:
            logger.warning(ex)
            if "uq_categories_name" in str(ex):
                raise exceptions.DuplicateCategoryName
        await redis.delete(f"sub-categories:{parent_category_id}")
    else:
        without_parent_query = sa.insert(Category).values(
            {
                Category.name: payload.name,
                Category.description: payload.description
            }
        )
        try:
            async with session.begin() as conn:
                await conn.execute(without_parent_query)
        except IntegrityError as ex:
            logger.warning(ex)
            if "uq_categories_name" in str(ex):
                raise exceptions.DuplicateCategoryName
        await redis.delete("root-categories")
    


async def all_categories(
        engine: AsyncEngine,
        limit: int,
        offset: int
) -> dict | None:
    parent_category_table_name = so.aliased(Category)
    parent_category_name = (parent_category_table_name.name).label("parent_name")
    query = (
        sa.select(
            Category.id,
            Category.name,
            Category.description,
            Category.is_active,
            parent_category_name
        )
        .select_from(Category)
        .join(
            parent_category_table_name,
            Category.parent_id==parent_category_table_name.id,
            isouter=True
        )
    ).order_by(Category.created_at.desc())
    return await paginate(engine=engine, query=query, limit=limit, offset=offset)


async def delete_category_by_id(
        session: async_sessionmaker[AsyncSession],
        category_id: CategoryId,
        redis: Redis
) -> None:
    query = sa.delete(Category).where(Category.id==category_id).returning(Category.id)
    try:
        async with session.begin() as conn:
            result = await conn.scalar(query)
            if result is None:
                raise exceptions.CategoryNotFound
    except exceptions.CategoryNotFound as ex:
        logger.warning(ex)
        raise exceptions.CategoryNotFound
    except IntegrityError as ex:
        logger.warning(ex)
    await redis.delete("root-categories")
    async for key in redis.scan_iter("sub-categories:*"):
        await redis.delete(key)


async def update_category_by_id(
        session: async_sessionmaker[AsyncSession],
        category_id: CategoryId,
        payload: schemas.Category,
        redis: Redis
) -> None:
    if payload.parent_category_name:
        parent_query = sa.select(Category.id).where(
            Category.name==payload.parent_category_name
        )
        try:
            async with session.begin() as conn:
                result: CategoryId | None = await conn.scalar(parent_query)
                if result is None:
                    raise exceptions.InvalidParentCategoryName
                updated_query = sa.update(Category).where(Category.id==category_id).values(
                    {
                        Category.name: payload.name,
                        Category.parent_id: result if payload.parent_category_name else None
                    }
                ).returning(Category.id)
                updated_result: CategoryId | None = await conn.scalar(updated_query)
                if updated_result is None:
                    raise exceptions.CategoryNotFound
            await redis.delete(f"sub-categories:{result}")
        except exceptions.InvalidParentCategoryName as ex:
            logger.warning(ex)
            raise exceptions.InvalidParentCategoryName
        except exceptions.CategoryNotFound as ex:
            logger.warning(ex)
            raise exceptions.CategoryNotFound
        except IntegrityError as ex:
            logger.warning(ex)
            if "uq_categories_name" in str(ex):
                raise exceptions.DuplicateCategoryName
    if not payload.parent_category_name:
        await redis.delete("root-categories")


async def activate_category(
        category_id: CategoryId,
        session: async_sessionmaker[AsyncSession],
        redis: Redis
) -> None:
    await redis.delete("root-categories")
    async for key in redis.scan_iter("sub-categories:*"):
        await redis.delete(key)
    query = sa.update(Category).where(Category.id==category_id).values(
        {
            Category.is_active: True
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except Exception as ex:
        logger.warning(ex)


async def deactivate_category(
        category_id: CategoryId,
        session: async_sessionmaker[AsyncSession],
        redis: Redis
) -> None:
    await redis.delete("root-categories")
    async for key in redis.scan_iter("sub-categories:*"):
        await redis.delete(key)
    query = sa.update(Category).where(Category.id==category_id).values(
        {
            Category.is_active: False
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except Exception as ex:
        logger.warning(ex)

# ==================== Attribute service ==================== #

async def create_attribute(
        session: async_sessionmaker[AsyncSession],
        payload: schemas.Attribute
) -> None:
    query = sa.insert(Attribute).values(
        {
            Attribute.name: payload.name
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)
        if "pk_attributes" in str(ex):
            raise exceptions.DuplicateAttributeName


async def list_attributes(
        engine: AsyncEngine,
        limit: int,
        offset: int,
        name__contain: str | None
) -> dict | None:
    query = sa.select(Attribute.name)
    if name__contain:
        query = query.where(Attribute.name.ilike(f"%{name__contain}%"))
    return await paginate(
        query=query, engine=engine, limit=limit, offset=offset
    )


async def delete_attribute(
        attribute_name: str,
        session: async_sessionmaker[AsyncSession],
) -> None:
    query = (
        sa.delete(Attribute)
        .where(Attribute.name==attribute_name)
        .returning(Attribute.name)
    )
    try:
        async with session.begin() as conn:
            result: str | None = await conn.scalar(query)
            if result is None:
                raise exceptions.AttributeNotFound
    except exceptions.AttributeNotFound as ex:
        logger.warning(ex)
        raise exceptions.AttributeNotFound
    except IntegrityError as ex:
        logger.warning(ex)


async def update_attribute(
        attribute_name: str,
        session: async_sessionmaker[AsyncSession],
        payload: schemas.Attribute
) -> None:
    query = (
        sa.update(Attribute)
        .where(Attribute.name==attribute_name)
        .values(
            {
                "name": payload.name
            }
        )
        .returning(Attribute.name)
    )
    try:
        async with session.begin() as conn:
            result: str | None = await conn.scalar(query)
            if result is None:
                raise exceptions.AttributeNotFound
    except exceptions.AttributeNotFound as ex:
        logger.warning(ex)
        raise exceptions.AttributeNotFound
    except IntegrityError as ex:
        if "pk_attributes" in str(ex):
            raise exceptions.DuplicateAttributeName

# ==================== CategoryAttributes service ==================== #

async def assign_category_attribute(
        session: async_sessionmaker[AsyncSession],
        attribute_name: str,
        category_id: CategoryId,
) -> None:
    query = sa.insert(CategoryAttribute).values(
            {
                CategoryAttribute.category_id: category_id,
                CategoryAttribute.attribute_name: attribute_name
            }
        )
    try:
        async with session.begin() as conn:
                await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)
        if "fk_categoryattributes_category_id_categories" in str(ex):
            raise exceptions.CategoryNotFound
        if "pk_categoryattributes" in str(ex):
            raise exceptions.CategoryAttributeUniqueTogether
        if "fk_categoryattributes_attribute_name_attributes" in str(ex):
            raise exceptions.AttributeNotFound


async def unassign_category_attribute(
        session: async_sessionmaker[AsyncSession],
        attribute_name: str,
        category_id: CategoryId,
) -> None:
    query = sa.delete(CategoryAttribute).where(
        sa.and_(
            CategoryAttribute.category_id==category_id,
            CategoryAttribute.attribute_name==attribute_name
        )
    ).returning(CategoryAttribute.category_id)
    try:
        async with session.begin() as conn:
            result: CategoryId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.UnassignedWentWrong
    except exceptions.UnassignedWentWrong as ex:
        logger.warning(ex)
        raise exceptions.UnassignedWentWrong
    except IntegrityError as ex:
        logger.warning(ex)

# ==================== Products service ==================== #

async def create_product(
        payload: schemas.ProductIn,
        images: list[UploadFile],
        session: async_sessionmaker[AsyncSession],
) -> None:
    image_unique_names = await validate_images_and_return_unique_image_names(images)

    category_query = sa.select(Category.id).where(
        Category.name==payload.category_name
    )
    brand_query = sa.select(Brand.id).where(
        Brand.name==payload.brand_name
    )
    try:
        async with session.begin() as conn:
            category_id: CategoryId | None = await conn.scalar(category_query)
            brand_id: BrandId | None = await conn.scalar(brand_query)
            if category_id is None:
                raise exceptions.CategoryNotFound
            if brand_id is None:
                raise exceptions.BrandNotFound
            product_query = sa.insert(Product).values(
                {
                    Product.serial_number: payload.serial_number,
                    Product.name: payload.name,
                    Product.description: payload.description,
                    Product.stock: payload.stock,
                    Product.price: payload.price,
                    Product.discount: payload.discount if payload.discount else None,
                    Product.expiry_discount: payload.expiry_discount if payload.expiry_discount else None,
                    Product.brand_id: brand_id,
                    Product.category_id: category_id
                }
            ).returning(Product.id)
            product_id: ProductId | None = await conn.scalar(product_query)
            image_query = sa.insert(ProductImage).values(
                [
                    {
                        ProductImage.url: image_name,
                        ProductImage.product_id: product_id
                    } for image_name in image_unique_names
                ]
            )
            if len(payload.attribute_values) >= 1:
                attribute_query = sa.insert(AttributeValue).values(
                    [
                        {
                            AttributeValue.value: attribute_value.value,
                            AttributeValue.attribute_name: attribute_value.attribute,
                            AttributeValue.product_id: product_id
                        } for attribute_value in payload.attribute_values
                    ]
                )
                await conn.execute(attribute_query)
            await conn.execute(image_query)
    except exceptions.CategoryNotFound as ex:
        logger.warning(ex)
        raise exceptions.CategoryNotFound
    except exceptions.BrandNotFound as ex:
        logger.warning(ex)
        raise exceptions.BrandNotFound
    except IntegrityError as ex:
        logger.warning(ex)
        if "uq_products_name" in str(ex):
            raise exceptions.DuplicateProductName
        if "uq_products_serial_number" in str(ex):
            raise exceptions.DuplicateProductSerialNumber
        if "fk_attributevalues_attribute_name_attributes" in str(ex):
            raise exceptions.AttributeNotFound
        if "fk_attributevalues_product_id_products" in str(ex):
            raise exceptions.ProductNotFound

    await asyncio.gather(*[
        upload_to_s3(file=image_file, unique_filename=image_unique_name)
        for image_unique_name, image_file in image_unique_names.items()
    ])


async def activate_product(
        session: async_sessionmaker[AsyncSession],
        product_id: ProductId
) -> None:
    query = sa.update(Product).where(Product.id==product_id).values(
        {
            Product.is_active: True
        }
    ).returning(Product.id)
    try:
        async with session.begin() as conn:
            result: ProductId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.ProductNotFound
    except exceptions.ProductNotFound as ex:
        logger.warning(ex)
        raise exceptions.ProductNotFound
    except IntegrityError as ex:
        logger.warning(ex)


async def deactivate_product(
        session: async_sessionmaker[AsyncSession],
        product_id: ProductId
) -> None:
    query = sa.update(Product).where(Product.id==product_id).values(
        {
            Product.is_active: False
        }
    ).returning(Product.id)
    try:
        async with session.begin() as conn:
            result: ProductId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.ProductNotFound
    except exceptions.ProductNotFound as ex:
        logger.warning(ex)
        raise exceptions.ProductNotFound
    except IntegrityError as ex:
        logger.warning(ex)


async def delete_product(
        session: async_sessionmaker[AsyncSession],
        product_id: ProductId
) -> None:
    image_query = sa.select(ProductImage.url).where(
        ProductImage.product_id==product_id
    )
    query = sa.delete(Product).where(Product.id==product_id)
    try:
        async with session.begin() as conn:
            result = list((await conn.scalars(image_query)).all())
            await conn.execute(query)
    except Exception as ex:
        logger.warning(ex)
    for image_name in result:
        await delete_from_s3(image_name)


async def list_products(
        filter_query: schemas.ProductQuerySearch,
        engine: AsyncEngine,
        limit: int,
        offset: int,
) -> dict | None:
    sub_query = sa.select(
        ProductImage.url,
        ProductImage.product_id
    ).distinct(ProductImage.product_id).subquery()
    query = (
        sa.select(
            Product.id,
            Product.serial_number,
            Product.category_id,
            Product.name,
            Product.stock,
            Product.price,
            Product.discount,
            Product.is_active,
            Category.name.label("category_name"),
            Brand.name.label("brand_name"),
            sub_query.c.url.label("image_url")
        )
        .select_from(Product)
        .join(Category, Product.category_id==Category.id)
        .join(Brand, Product.brand_id==Brand.id)
        .join(sub_query, Product.id==sub_query.c.product_id)
    )

    if filter_query.brand__exact:
        query = query.where(
            Brand.name==filter_query.brand__exact
        )

    if filter_query.category__exact:
        categories_cte = sa.select(
            Category.id,
            Category.name,
            Category.parent_id,
            sa.literal(0).label("level")
        ).where(Category.name==filter_query.category__exact).cte(recursive=True)

        category_alias = sa.alias(Category) # type: ignore

        recursive_query = sa.select(
            category_alias.c.id,
            category_alias.c.name,
            category_alias.c.parent_id,
            (categories_cte.c.level + 1).label("level")
        ).join(categories_cte, categories_cte.c.id==category_alias.c.parent_id)

        categories_cte = categories_cte.union(recursive_query)
        query = query.join(
            categories_cte, Product.category_id==categories_cte.c.id
        ).order_by(
            categories_cte.c.level
        )

    if filter_query.name__contain:
        query = query.where(
            Product.name.ilike(f"%{filter_query.name__contain}%")
        )

    query = query.order_by(Product.created_at.desc())

    return await paginate(
        engine=engine, query=query, limit=limit, offset=offset
    )


async def product_detail(
        session: async_sessionmaker[AsyncSession],
        product_serial: SerialNumber,
) -> AdminProductDetailResponse:
    query = (
        sa.select(
            Product.id,
            Product.serial_number,
            Product.name,
            Product.stock,
            Product.price,
            Product.discount,
            Product.is_active,
            Product.description,
            Product.expiry_discount,
            Category.name.label("category_name"),
            Brand.name.label("brand_name"),
            AttributeValue.attribute_name.label("attribute"),
            AttributeValue.value,
            ProductImage.url.label("image_urls")
        )
        .select_from(Product)
        .join(Category, Product.category_id==Category.id)
        .join(Brand, Product.brand_id==Brand.id)
        .join(ProductImage, Product.id==ProductImage.product_id, isouter=True)
        .join(AttributeValue, Product.id==AttributeValue.product_id, isouter=True)
        .where(Product.serial_number==product_serial)
    )
    try:
        async with session.begin() as conn:
            result = (await conn.execute(query)).all()
    except Exception as ex:
        logger.warning(ex)

    if len(result) == 0:
        raise exceptions.ProductNotFound
    attribute_values = dict()
    for p in result:
        if p.attribute not in attribute_values:
            attribute_values[p.attribute] = p.value
    return {
        "id": result[0].id,
        "serial_number": result[0].serial_number,
        "is_active": result[0].is_active,
        "name": result[0].name,
        "stock": result[0].stock,
        "price": result[0].price,
        "discount": result[0].discount,
        "description": result[0].description,
        "expiry_discount": result[0].expiry_discount,
        "category_name": result[0].category_name,
        "brand_name": result[0].brand_name,
        "image_urls": set([p.image_urls for p in result]),
        "attribute_values": attribute_values
    }

# ==================== Comment service ==================== #

async def delete_comment(
        comment_id: CommentId,
        session: async_sessionmaker[AsyncSession],
) -> None:
    query = sa.delete(Comment).where(Comment.id==comment_id)
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)

# ==================== Guaranty service ==================== #

async def add_guaranties(
        redis: Redis,
        file: UploadFile
) -> None:
    unique_file_name = await create_unique_excel_name(file=file)
    await upload_to_s3(file=file.file, unique_filename=unique_file_name)
    await redis.set(name=f"file:{unique_file_name}", value=1, ex=1)

# ==================== Extract excel data and insert them to db ==================== #

async def process_excel_data(
        file: BytesIO,
        session: async_sessionmaker[AsyncSession],
        batch_size: int = 500
) -> None:
    workbook = load_workbook(file)
    sheet = workbook.active

    all_rows: list[ExcelEntityTypes] = [
        {
            "product_serial_number":row[0],
            "guaranty_serial": row[1],
            "product_name": row[2],
            "guaranty_days": row[3],
            "produced_at": row[4].replace(" ق.ظ", "").replace(" ب.ظ", "")
        } for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

    def chunks(data: list, chunk_size: int):
        for chunk in range(0, len(data), chunk_size):
            yield data[chunk: chunk + chunk_size]

    try:
        async with session.begin() as conn:
            for batch in chunks(all_rows, batch_size):
                query = postgres_upsert(Guaranty).values(batch)
                query = query.on_conflict_do_nothing()
                await conn.execute(query)
            # TODO: Send a message for admin user if data processed successfully here.
    except IntegrityError as ex:
        logger.warning(ex)

# ==================== Ticket service ==================== #

async def list_tickets(
        engine: AsyncEngine,
        limit: int,
        offset: int
) -> dict | None:
    query = sa.select(
        Ticket.id,
        Ticket.name,
        Ticket.product_serial,
        Ticket.phone_number,
        Ticket.guaranty_rating,
        Ticket.repairs_rating,
        Ticket.notification_rating,
        Ticket.personal_behavior_rating,
        Ticket.services_rating,
        Ticket.smart_process_rating,
        Ticket.criticism,
        Ticket.call_request
    ).order_by(Ticket.created_at.desc())
    return await paginate(
        engine=engine, query=query, limit=limit, offset=offset
    )


async def delete_ticket(
        session: async_sessionmaker[AsyncSession],
        ticket_id: TicketId
) -> None:
    query = sa.delete(Ticket).where(
        Ticket.id==ticket_id
    ).returning(Ticket.id)
    try:
        async with session.begin() as conn:
            result: TicketId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.TicketNotFound
    except exceptions.TicketNotFound as ex:
        logger.warning(ex)
        raise exceptions.TicketNotFound

# ==================== Article service ==================== #

async def create_article(
        session: async_sessionmaker[AsyncSession],
        payload: schemas.ArticleIn,
        images: list[UploadFile],
) -> None:
    image_unique_names = await validate_images_and_return_unique_image_names(images)
    query = sa.insert(Article).values(
        {
            Article.title: payload.title,
            Article.description: payload.description
        }
    ).returning(Article.id)
    try:
        async with session.begin() as conn:
            result: ArticleId | None = await conn.scalar(query)
            image_query = sa.insert(ArticleImage).values(
                [
                    {
                        ArticleImage.article_id: result,
                        ArticleImage.url: image_name
                    } for image_name in image_unique_names
                ]
            )
            await conn.execute(image_query)
    except IntegrityError as ex:
        logger.warning(ex)
        if "uq_articles_title" in str(ex):
            raise exceptions.DuplicateArticleTitle

    await asyncio.gather(*[
        upload_to_s3(file=image_file, unique_filename=image_unique_name)
        for image_unique_name, image_file in image_unique_names.items()
    ])


async def delete_article(
        session: async_sessionmaker[AsyncSession],
        article_id: ArticleId
) -> None:
    image_query = sa.select(ArticleImage.url).where(
        ArticleImage.article_id==article_id
    )
    query = sa.delete(Article).where(Article.id==article_id)
    try:
        async with session.begin() as conn:
            result = list((await conn.scalars(image_query)).all())
            await conn.execute(query)
    except Exception as ex:
        logger.warning(ex)
    if result:
        for image_name in result:
            await delete_from_s3(image_name)

# ==================== Tag service ==================== #

async def create_tag(
        session: async_sessionmaker[AsyncSession],
        payload: schemas.Tag
) -> None:
    query = sa.insert(Tag).values(
        {
            Tag.name: payload.name
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)
        if "pk_tags" in str(ex):
            raise exceptions.DuplicateTagName


async def list_tags(
        engine: AsyncEngine,
        limit: int,
        offset: int,
        name__contain: str | None
) -> dict | None:
    query = sa.select(Tag.name)
    if name__contain:
        query = query.where(Tag.name.ilike(f"%{name__contain}%"))
    return await paginate(
        query=query, engine=engine, limit=limit, offset=offset
    )


async def delete_tag(
        tag_name: str,
        session: async_sessionmaker[AsyncSession],
) -> None:
    query = (
        sa.delete(Tag)
        .where(Tag.name==tag_name)
        .returning(Tag.name)
    )
    try:
        async with session.begin() as conn:
            result: str | None = await conn.scalar(query)
            if result is None:
                raise exceptions.TagNotFound

    except exceptions.TagNotFound as ex:
        logger.warning(ex)
        raise exceptions.TagNotFound

    except IntegrityError as ex:
        logger.warning(ex)


async def update_tag(
        tag_name: str,
        session: async_sessionmaker[AsyncSession],
        new_name: str
) -> None:
    query = (
        sa.update(Tag)
        .values(
        {
            Tag.name: new_name
        }
        )
        .where(Tag.name==tag_name)
        .returning(Tag.name)
    )
    try:
        async with session.begin() as conn:
            result: str | None = await conn.scalar(query)
            if result is None:
                raise exceptions.TagNotFound

    except exceptions.TagNotFound as ex:
        logger.warning(ex)
        raise exceptions.TagNotFound

    except IntegrityError as ex:
        logger.warning(ex)

# ==================== ArticleTag service ==================== #

async def assign_tags_to_article(
        article_id: ArticleId,
        session: async_sessionmaker[AsyncSession],
        tag_name: str
) -> None:
    tag_query = sa.select(Tag.name).where(Tag.name==tag_name)
    try:
        async with session.begin() as conn:
            tag_result: str | None = await conn.scalar(tag_query)
            if tag_result is None:
                raise exceptions.TagNotFound
            query = sa.insert(ArticleTag).values(
                {
                    ArticleTag.article_id: article_id,
                    ArticleTag.tag_name: tag_result
                }
            )
            await conn.execute(query)

    except exceptions.TagNotFound as ex:
        logger.warning(ex)
        raise exceptions.TagNotFound

    except IntegrityError as ex:
        logger.warning(ex)

        if "pk_article_tags" in str(ex):
            raise exceptions.DuplicateArticleTagPk

        if "fk_article_tags_article_id_articles" in str(ex):
            raise ArticleNotFound


async def unassign_tags_to_article(
        article_id: ArticleId,
        session: async_sessionmaker[AsyncSession],
        tag_name: str
) -> None:
    tag_query = sa.select(Tag.name).where(Tag.name==tag_name)
    try:
        async with session.begin() as conn:
            tag_result: str | None = await conn.scalar(tag_query)
            if tag_result is None:
                raise exceptions.TagNotFound
            query = sa.delete(ArticleTag).where(
                sa.and_(
                    ArticleTag.article_id==article_id,
                    ArticleTag.tag_name==tag_result
                )
            )
            await conn.execute(query)

    except exceptions.TagNotFound as ex:
        logger.warning(ex)
        raise exceptions.TagNotFound

    except IntegrityError as ex:
        logger.warning(ex)

# ==================== Glossary service ==================== #

async def create_glossary(
        article_id: ArticleId,
        session: async_sessionmaker[AsyncSession],
        payload: schemas.GlossaryIn
) -> None:
    query = sa.insert(GlossaryTerm).values(
        {
            GlossaryTerm.term: payload.term,
            GlossaryTerm.definition: payload.definition,
            GlossaryTerm.article_id: article_id
        }
    )
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)
        if "uq_glossary_terms_term" in str(ex):
            raise exceptions.UniqueConstraintGlossaryTerms
        if "fk_glossary_terms_article_id_articles" in str(ex):
            raise ArticleNotFound


async def delete_glossary(
        session: async_sessionmaker[AsyncSession],
        glossary_id: GlossaryId
) -> None:
    query = (
        sa.delete(GlossaryTerm)
        .where(GlossaryTerm.id==glossary_id)
        .returning(GlossaryTerm.id)
    )
    try:
        async with session.begin() as conn:
            result: GlossaryId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.GlossaryNotFound
    except exceptions.GlossaryNotFound as ex:
        logger.warning(ex)
        raise exceptions.GlossaryNotFound
    except IntegrityError as ex:
        logger.warning(ex)


async def update_glossary(
        session: async_sessionmaker[AsyncSession],
        glossary_id: GlossaryId,
        payload: schemas.GlossaryIn
) -> None:
    query = (
        sa.update(GlossaryTerm)
        .values(
            {
                GlossaryTerm.term: payload.term,
                GlossaryTerm.definition: payload.definition
            }
        )
        .where(GlossaryTerm.id==glossary_id)
        .returning(GlossaryTerm.id)
    )
    try:
        async with session.begin() as conn:
            result: GlossaryId | None = await conn.scalar(query)
            if result is None:
                raise exceptions.GlossaryNotFound
    except exceptions.GlossaryNotFound as ex:
        logger.warning(ex)
        raise exceptions.GlossaryNotFound
    except IntegrityError as ex:
        logger.warning(ex)
        if "uq_glossary_terms_term" in str(ex):
            raise exceptions.UniqueConstraintGlossary

# ==================== ArticleComment service ==================== #

async def delete_article_comment(
        article_comment_id: ArticleCommentId,
        session: async_sessionmaker[AsyncSession],
) -> None:
    query = sa.delete(ArticleComment).where(ArticleComment.id==article_comment_id)
    try:
        async with session.begin() as conn:
            await conn.execute(query)
    except IntegrityError as ex:
        logger.warning(ex)
